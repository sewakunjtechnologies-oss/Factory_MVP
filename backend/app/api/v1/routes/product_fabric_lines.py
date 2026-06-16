from __future__ import annotations

import io
from decimal import Decimal, InvalidOperation
from typing import Any, List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from typing_extensions import Annotated

from app.core.database import get_db
from app.core.security import require_owner, require_owner_or_manager
from app.models.product import Product
from app.models.product_fabric_line import ProductFabricLine
from app.models.purchase_order import PurchaseOrder
from app.models.user import User
from app.schemas.product_fabric_line import (
    ProductFabricLineCreate,
    ProductFabricLineRead,
    ProductFabricLineUpdate,
)
from app.services.fabric_planning import build_or_refresh_fabric_plan

router = APIRouter()


@router.get("", response_model=List[ProductFabricLineRead])
async def list_lines(
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(require_owner_or_manager)],
    product_id: Optional[UUID] = Query(default=None),
) -> List[ProductFabricLineRead]:
    stmt = select(ProductFabricLine)
    if product_id is not None:
        stmt = stmt.where(ProductFabricLine.product_id == product_id)
    stmt = stmt.order_by(ProductFabricLine.fabric_code.asc())
    result = await db.execute(stmt)
    return result.scalars().all()


@router.post("", response_model=ProductFabricLineRead, status_code=201)
async def create_line(
    payload: ProductFabricLineCreate,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(require_owner)],
) -> ProductFabricLineRead:
    product = await db.get(Product, payload.product_id)
    if product is None:
        raise HTTPException(status_code=404, detail="Product not found")
    line = ProductFabricLine(**payload.model_dump())
    db.add(line)
    await db.commit()
    await db.refresh(line)
    return line


@router.patch("/{line_id}", response_model=ProductFabricLineRead)
async def update_line(
    line_id: UUID,
    payload: ProductFabricLineUpdate,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(require_owner_or_manager)],
) -> ProductFabricLineRead:
    line = await db.get(ProductFabricLine, line_id)
    if line is None:
        raise HTTPException(status_code=404, detail="Fabric line not found")
    updates = payload.model_dump(exclude_unset=True)
    for field, value in updates.items():
        setattr(line, field, value)
    await _refresh_matching_po_plans(db, line)
    await db.commit()
    await db.refresh(line)
    return line


@router.delete("/{line_id}", status_code=204, response_class=Response)
async def delete_line(
    line_id: UUID,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(require_owner)],
) -> Response:
    line = await db.get(ProductFabricLine, line_id)
    if line is None:
        raise HTTPException(status_code=404, detail="Fabric line not found")
    product_id = line.product_id
    fabric_code = line.fabric_code
    await db.delete(line)
    await db.flush()
    await _refresh_matching_po_plans_by_key(db, product_id, fabric_code)
    await db.commit()
    return Response(status_code=204)


async def _refresh_matching_po_plans(db: AsyncSession, line: ProductFabricLine) -> None:
    await _refresh_matching_po_plans_by_key(db, line.product_id, line.fabric_code)


async def _refresh_matching_po_plans_by_key(db: AsyncSession, product_id: UUID, fabric_code: str) -> None:
    result = await db.execute(
        select(PurchaseOrder).where(
            PurchaseOrder.product_id == product_id,
            PurchaseOrder.design_code_snapshot == fabric_code,
        )
    )
    for po in result.scalars().all():
        await build_or_refresh_fabric_plan(db, po)


# ---------------------------------------------------------------------------
# Excel import / export
# ---------------------------------------------------------------------------

_EXCEL_COLUMNS = [
    "category",
    "fabric_code",
    "pieces",
    "pieces_in_stock",
    "pieces_short",
    "stock_status",
    "per_piece_meters",
    "stock_meters",
    "pieces_per_bale",
    "bale_size_cbm",
    "bale_weight_kg",
]

_VALID_STOCK_STATUS = {"extra", "in_stock", "ok", "nil", "short", "unknown"}


@router.get("/export-xlsx")
async def export_xlsx(
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(require_owner_or_manager)],
) -> StreamingResponse:
    """Download the current inventory as an Excel file.

    The owner edits the numbers in Excel, then re-uploads via import-xlsx to
    update the database in one shot.
    """
    from openpyxl import Workbook

    stmt = (
        select(ProductFabricLine, Product)
        .join(Product, Product.id == ProductFabricLine.product_id)
        .order_by(Product.product_name.asc(), ProductFabricLine.fabric_code.asc())
    )
    rows = (await db.execute(stmt)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Fabric Inventory"
    ws.append(_EXCEL_COLUMNS)
    for line, product in rows:
        ws.append([
            product.product_name,
            line.fabric_code,
            int(line.pieces or 0),
            int(line.pieces_in_stock or 0),
            int(line.pieces_short or 0),
            line.stock_status or "unknown",
            float(line.per_piece_meters or 0),
            float(line.stock_meters or 0),
            int(line.pieces_per_bale or 0),
            float(line.bale_size_cbm or 0),
            float(line.bale_weight_kg or 0),
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="fabric_inventory.xlsx"'},
    )


def _normalize_status(raw: Any) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip().lower().replace(" ", "_").replace("-", "_")
    # Friendly synonyms the owner might write in the sheet.
    synonyms = {
        "ok": "ok",
        "extra": "extra",
        "in_stock": "in_stock",
        "instock": "in_stock",
        "stock": "in_stock",
        "nil": "nil",
        "zero": "nil",
        "short": "short",
        "deficit": "short",
        "shortage": "short",
        "unknown": "unknown",
        "": "unknown",
    }
    return synonyms.get(s, s if s in _VALID_STOCK_STATUS else None)


def _coerce_int(raw: Any) -> int | None:
    if raw is None or raw == "":
        return None
    try:
        return int(float(raw))
    except (TypeError, ValueError):
        return None


def _coerce_decimal(raw: Any) -> Decimal | None:
    if raw is None or raw == "":
        return None
    try:
        return Decimal(str(raw))
    except (TypeError, ValueError, InvalidOperation):
        return None


@router.post("/import-xlsx")
async def import_xlsx(
    file: Annotated[UploadFile, File(description="Excel sheet with columns: category, fabric_code, pieces, pieces_in_stock, pieces_short, stock_status, per_piece_meters, stock_meters")],
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(require_owner_or_manager)],
    dry_run: bool = Query(default=False, description="If true, report what would change but do not commit."),
) -> dict:
    """Bulk-update inventory from an Excel file.

    Rows are matched by (category, fabric_code) — both compared case-insensitively.
    Unmatched rows are reported in the response so the owner can fix typos.
    Empty cells leave the existing value untouched.
    """
    from openpyxl import load_workbook

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Upload an .xlsx file.")

    raw = await file.read()
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Could not read the Excel file: {e}") from e

    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="The Excel file has no active sheet.")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="The sheet is empty.")

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    if "category" not in header or "fabric_code" not in header:
        raise HTTPException(
            status_code=400,
            detail=f"Header row must include 'category' and 'fabric_code'. Got: {header}",
        )

    def col(name: str) -> int | None:
        return header.index(name) if name in header else None

    idx = {name: col(name) for name in _EXCEL_COLUMNS}

    # Preload all lines once so we don't hit the DB per row.
    all_lines = (await db.execute(
        select(ProductFabricLine, Product).join(Product, Product.id == ProductFabricLine.product_id)
    )).all()
    by_key: dict[tuple[str, str], ProductFabricLine] = {}
    for line, product in all_lines:
        by_key[(product.product_name.strip().lower(), line.fabric_code.strip().lower())] = line

    updated = 0
    skipped: list[dict[str, Any]] = []
    changes: list[dict[str, Any]] = []

    for row_number, raw_row in enumerate(rows[1:], start=2):
        if all(v is None or str(v).strip() == "" for v in raw_row):
            continue
        category = str(raw_row[idx["category"]] or "").strip()
        fabric_code = str(raw_row[idx["fabric_code"]] or "").strip()
        if not category or not fabric_code:
            skipped.append({"row": row_number, "reason": "missing category or fabric_code"})
            continue

        line = by_key.get((category.lower(), fabric_code.lower()))
        if line is None:
            skipped.append({"row": row_number, "reason": f"no fabric line for {category} / {fabric_code}"})
            continue

        change: dict[str, Any] = {"row": row_number, "category": category, "fabric_code": fabric_code, "fields": {}}

        def maybe_set_int(field: str) -> None:
            if idx[field] is None:
                return
            raw_value = raw_row[idx[field]]
            new_value = _coerce_int(raw_value)
            if new_value is None and raw_value not in (None, ""):
                skipped.append({"row": row_number, "reason": f"{field}={raw_value!r} is not a number"})
                return
            if new_value is None:
                return
            old = int(getattr(line, field) or 0)
            if old != new_value:
                change["fields"][field] = {"from": old, "to": new_value}
                if not dry_run:
                    setattr(line, field, new_value)

        def maybe_set_decimal(field: str) -> None:
            if idx[field] is None:
                return
            raw_value = raw_row[idx[field]]
            new_value = _coerce_decimal(raw_value)
            if new_value is None and raw_value not in (None, ""):
                skipped.append({"row": row_number, "reason": f"{field}={raw_value!r} is not a number"})
                return
            if new_value is None:
                return
            old = Decimal(str(getattr(line, field) or 0))
            if old != new_value:
                change["fields"][field] = {"from": float(old), "to": float(new_value)}
                if not dry_run:
                    setattr(line, field, new_value)

        maybe_set_int("pieces")
        maybe_set_int("pieces_in_stock")
        maybe_set_int("pieces_short")
        maybe_set_int("pieces_per_bale")
        maybe_set_decimal("per_piece_meters")
        maybe_set_decimal("stock_meters")
        maybe_set_decimal("bale_size_cbm")
        maybe_set_decimal("bale_weight_kg")

        if idx["stock_status"] is not None:
            raw_status = raw_row[idx["stock_status"]]
            new_status = _normalize_status(raw_status)
            if new_status is None and raw_status not in (None, ""):
                skipped.append({"row": row_number, "reason": f"stock_status={raw_status!r} is not recognized"})
            elif new_status and new_status != line.stock_status:
                change["fields"]["stock_status"] = {"from": line.stock_status, "to": new_status}
                if not dry_run:
                    line.stock_status = new_status

        if change["fields"]:
            updated += 1
            changes.append(change)

    if not dry_run:
        await db.commit()

    return {
        "dry_run": dry_run,
        "updated": updated,
        "skipped": skipped,
        "changes": changes[:50],  # cap the response size; full audit is in the DB after commit
    }
