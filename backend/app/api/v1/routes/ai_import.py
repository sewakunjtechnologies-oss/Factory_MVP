"""AI-driven Excel import.

Two endpoints:

  POST /api/v1/ai-import/analyze   — upload xlsx, get Gemini's structured plan +
                                     a preview of what would change.
  POST /api/v1/ai-import/commit    — upload xlsx + the plan; apply changes
                                     atomically.

The analyze step is the only one that calls Gemini. Commit is pure execution
against Postgres so it stays predictable and auditable.
"""

from __future__ import annotations

import io
import json
from datetime import date
from decimal import Decimal
from typing import Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from typing_extensions import Annotated

from app.core.database import get_db
from app.core.security import require_owner_or_manager
from app.models.fabric_meter_receipt import FabricMeterReceipt
from app.models.pieces_receipt import PiecesReceipt
from app.models.product import Product
from app.models.product_fabric_line import ProductFabricLine
from app.models.user import User
from app.services.ai_excel_import import (
    TABLE_CATALOG,
    ImportPlan,
    analyze_with_gemini,
    coerce_decimal,
    coerce_int,
    coerce_status,
    coerce_str,
    read_excel_sample,
)

router = APIRouter()

_MAX_FILE_BYTES = 5 * 1024 * 1024  # 5 MB — anything larger isn't a fabric stock sheet


async def _read_uploaded_xlsx(file: UploadFile) -> bytes:
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Upload an .xlsx file.")
    raw = await file.read()
    if len(raw) > _MAX_FILE_BYTES:
        raise HTTPException(status_code=400, detail=f"File is larger than {_MAX_FILE_BYTES // (1024 * 1024)} MB.")
    if not raw:
        raise HTTPException(status_code=400, detail="The uploaded file is empty.")
    return raw


@router.post("/analyze")
async def analyze(
    file: Annotated[UploadFile, File(description="Excel file (.xlsx) to analyze.")],
    _: Annotated[User, Depends(require_owner_or_manager)],
) -> dict[str, Any]:
    """Look at the headers + first 10 rows and ask Gemini what this is."""
    raw = await _read_uploaded_xlsx(file)

    try:
        sample = read_excel_sample(raw, sample_rows=10)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e

    try:
        plan = await analyze_with_gemini(sample)
    except Exception as e:  # noqa: BLE001
        # Surface a friendly error instead of a 500; the UI can show "AI is offline".
        raise HTTPException(status_code=502, detail=f"AI analysis failed: {e}") from e

    return {
        "plan": {
            "target_table": plan.target_table,
            "column_mapping": plan.column_mapping,
            "match_columns": plan.match_columns,
            "action": plan.action,
            "confidence": plan.confidence,
            "reasoning": plan.reasoning,
            "warnings": plan.warnings,
        },
        "table_info": TABLE_CATALOG[plan.target_table],
        "sample": {
            "sheet": sample.sheet_name,
            "headers": sample.headers,
            "first_rows": sample.rows,
            "approx_total_rows": sample.total_rows,
        },
    }


@router.post("/commit")
async def commit(
    file: Annotated[UploadFile, File(description="The same Excel file analyzed earlier.")],
    plan_json: Annotated[str, Form(description="The plan JSON returned by /analyze.")],
    db: Annotated[AsyncSession, Depends(get_db)],
    user: Annotated[User, Depends(require_owner_or_manager)],
) -> dict[str, Any]:
    """Apply the plan against the DB. Each-row failures are reported, not fatal."""
    raw = await _read_uploaded_xlsx(file)

    try:
        plan_data = json.loads(plan_json)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"Invalid plan JSON: {e}") from e

    target = plan_data.get("target_table")
    if target not in TABLE_CATALOG:
        raise HTTPException(status_code=400, detail=f"Unknown target_table: {target}")

    column_mapping: dict[str, str] = plan_data.get("column_mapping") or {}
    match_columns: dict[str, str] = plan_data.get("match_columns") or {}

    # Re-parse the whole workbook (not just the sample) — we need every row to commit.
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Could not read the Excel file: {e}") from e
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="The Excel file has no active sheet.")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="The sheet is empty.") from None
    headers = [str(c).strip() if c is not None else "" for c in header_row]
    header_index = {h.lower(): i for i, h in enumerate(headers) if h}

    def col_idx(excel_header: str) -> int | None:
        if not excel_header:
            return None
        return header_index.get(excel_header.strip().lower())

    # Preload every product_fabric_line keyed by (category_lower, fabric_code_lower).
    all_lines = (await db.execute(
        select(ProductFabricLine, Product).join(Product, Product.id == ProductFabricLine.product_id)
    )).all()
    line_by_key: dict[tuple[str, str], ProductFabricLine] = {}
    for line, product in all_lines:
        key = (product.product_name.strip().lower(), line.fabric_code.strip().lower())
        line_by_key[key] = line

    category_col = col_idx(match_columns.get("category", ""))
    fabric_col = col_idx(match_columns.get("fabric_code", ""))

    skipped: list[dict[str, Any]] = []
    applied: list[dict[str, Any]] = []
    applied_count = 0

    for row_number, raw_row in enumerate(rows_iter, start=2):
        if all(v is None or str(v).strip() == "" for v in raw_row):
            continue

        if category_col is None or fabric_col is None:
            skipped.append({"row": row_number, "reason": "plan did not specify which Excel columns hold category + fabric_code"})
            continue

        category = coerce_str(raw_row[category_col] if category_col < len(raw_row) else None)
        fabric_code = coerce_str(raw_row[fabric_col] if fabric_col < len(raw_row) else None)
        if not category or not fabric_code:
            skipped.append({"row": row_number, "reason": "missing category or fabric_code"})
            continue

        line = line_by_key.get((category.lower(), fabric_code.lower()))
        if line is None:
            skipped.append({"row": row_number, "reason": f"no fabric line for {category} / {fabric_code}"})
            continue

        try:
            if target == "product_fabric_lines":
                change = _apply_product_fabric_line_update(line, raw_row, column_mapping, col_idx)
            elif target == "pieces_receipts":
                change = _apply_pieces_receipt_insert(db, line, raw_row, column_mapping, col_idx, user_id=user.id)
            elif target == "fabric_meter_receipts":
                change = _apply_fabric_meter_receipt_insert(db, line, raw_row, column_mapping, col_idx, user_id=user.id)
            else:
                skipped.append({"row": row_number, "reason": f"target {target} is not implemented"})
                continue
        except ValueError as e:
            skipped.append({"row": row_number, "reason": str(e)})
            continue

        if change:
            applied.append({
                "row": row_number,
                "category": category,
                "fabric_code": fabric_code,
                **change,
            })
            applied_count += 1

    if applied_count > 0:
        await db.commit()

    return {
        "target_table": target,
        "applied": applied_count,
        "skipped": skipped,
        "changes": applied[:100],
    }


# ---------------------------------------------------------------------------
# Per-row appliers — small, single-purpose, never commit themselves.
# ---------------------------------------------------------------------------


def _apply_product_fabric_line_update(
    line: ProductFabricLine,
    raw_row: tuple[Any, ...],
    column_mapping: dict[str, str],
    col_idx,
) -> dict[str, Any] | None:
    """Update only the columns the user actually provided. Empty cells are ignored."""
    fields: dict[str, Any] = {}

    def get(excel_header: str) -> Any:
        idx = col_idx(excel_header)
        if idx is None or idx >= len(raw_row):
            return None
        return raw_row[idx]

    for db_col, excel_col in column_mapping.items():
        raw_value = get(excel_col)
        if raw_value is None or (isinstance(raw_value, str) and raw_value.strip() == ""):
            continue

        if db_col in ("pieces", "pieces_in_stock", "pieces_short"):
            v = coerce_int(raw_value)
            if v is None:
                raise ValueError(f"{db_col}={raw_value!r} is not a number")
            old = int(getattr(line, db_col) or 0)
            if old != v:
                fields[db_col] = {"from": old, "to": v}
                setattr(line, db_col, v)
        elif db_col in ("per_piece_meters", "stock_meters"):
            v = coerce_decimal(raw_value)
            if v is None:
                raise ValueError(f"{db_col}={raw_value!r} is not a number")
            old = Decimal(str(getattr(line, db_col) or 0))
            if old != v:
                fields[db_col] = {"from": float(old), "to": float(v)}
                setattr(line, db_col, v)
        elif db_col == "stock_status":
            v = coerce_status(raw_value)
            if v is None:
                raise ValueError(f"stock_status={raw_value!r} is not recognized")
            if line.stock_status != v:
                fields[db_col] = {"from": line.stock_status, "to": v}
                line.stock_status = v
        # Silently ignore mappings to columns we don't recognize for this table.

    return {"fields": fields} if fields else None


def _apply_pieces_receipt_insert(
    db: AsyncSession,
    line: ProductFabricLine,
    raw_row: tuple[Any, ...],
    column_mapping: dict[str, str],
    col_idx,
    *,
    user_id,
) -> dict[str, Any] | None:
    def get(excel_header: str) -> Any:
        idx = col_idx(excel_header)
        if idx is None or idx >= len(raw_row):
            return None
        return raw_row[idx]

    pieces = coerce_int(get(column_mapping.get("pieces", "")))
    if not pieces or pieces <= 0:
        raise ValueError("pieces is missing or not positive")

    received_at_raw = get(column_mapping.get("received_at", ""))
    received_at = _coerce_date(received_at_raw) or date.today()

    mill_name = coerce_str(get(column_mapping.get("mill_name", "")))
    notes = coerce_str(get(column_mapping.get("notes", "")))

    receipt = PiecesReceipt(
        product_fabric_line_id=line.id,
        pieces=pieces,
        received_at=received_at,
        mill_name=mill_name,
        notes=notes,
        created_by=user_id,
    )
    db.add(receipt)
    line.pieces_in_stock = (line.pieces_in_stock or 0) + pieces
    return {"inserted": "pieces_receipt", "pieces": pieces, "received_at": received_at.isoformat()}


def _apply_fabric_meter_receipt_insert(
    db: AsyncSession,
    line: ProductFabricLine,
    raw_row: tuple[Any, ...],
    column_mapping: dict[str, str],
    col_idx,
    *,
    user_id,
) -> dict[str, Any] | None:
    def get(excel_header: str) -> Any:
        idx = col_idx(excel_header)
        if idx is None or idx >= len(raw_row):
            return None
        return raw_row[idx]

    meters = coerce_decimal(get(column_mapping.get("meters", "")))
    if not meters or meters <= 0:
        raise ValueError("meters is missing or not positive")

    received_at_raw = get(column_mapping.get("received_at", ""))
    received_at = _coerce_date(received_at_raw) or date.today()

    mill_name = coerce_str(get(column_mapping.get("mill_name", "")))
    notes = coerce_str(get(column_mapping.get("notes", "")))

    receipt = FabricMeterReceipt(
        product_fabric_line_id=line.id,
        meters=meters,
        received_at=received_at,
        mill_name=mill_name,
        notes=notes,
        created_by=user_id,
    )
    db.add(receipt)
    line.stock_meters = Decimal(line.stock_meters or 0) + meters
    return {"inserted": "fabric_meter_receipt", "meters": float(meters), "received_at": received_at.isoformat()}


def _coerce_date(raw: Any) -> date | None:
    """Accept a python date, datetime, or an ISO/dd-mm-yyyy/dd/mm/yyyy string."""
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    # Try ISO first.
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        pass
    # Then dd-mm-yyyy or dd/mm/yyyy (common in Indian factory sheets).
    for sep in ("-", "/"):
        parts = s.split(sep)
        if len(parts) == 3:
            try:
                d, m, y = (int(parts[0]), int(parts[1]), int(parts[2]))
                if y < 100:
                    y += 2000
                return date(y, m, d)
            except ValueError:
                continue
    return None
