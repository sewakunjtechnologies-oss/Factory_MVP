"""AI-driven Excel importer.

Given an uploaded Excel file, ask Gemini to figure out which DB table the rows
belong to and how each Excel column maps to a writable DB column. The model
returns a structured plan; the route then applies the plan in a separate step
so the owner can review what's going to change before commit.
"""

from __future__ import annotations

import io
import json
import logging
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any, Literal

from google.genai import types

from app.services.voice.client import get_client, get_model

logger = logging.getLogger(__name__)


TargetTable = Literal["product_fabric_lines", "pieces_receipts", "fabric_meter_receipts"]
ActionKind = Literal["upsert", "insert", "update_only"]


# Catalog of tables the AI is allowed to write to, with the matching keys and
# writable columns. Kept right here so the LLM prompt and the executor stay in
# sync — adding a new target table is one edit.
TABLE_CATALOG: dict[str, dict[str, Any]] = {
    "product_fabric_lines": {
        "description": (
            "Per-category fabric inventory. One row per (category, fabric variant). "
            "Use this when the Excel looks like a stock sheet — columns like "
            "fabric name, pieces, in stock, short, meters."
        ),
        "match_key": ["category", "fabric_code"],
        "writable_columns": {
            "pieces": "int target piece count",
            "pieces_in_stock": "int finished pieces in warehouse",
            "pieces_short": "int pieces we owe (deficit)",
            "stock_status": "one of: extra, in_stock, ok, nil, short, unknown",
            "per_piece_meters": "decimal — meters of fabric per piece",
            "stock_meters": "decimal — fabric on hand in meters",
            "pieces_per_bale": "int — how many finished pieces fit in one bale",
            "bale_size_cbm": "decimal — volume of one bale in cubic meters (m³)",
            "bale_weight_kg": "decimal — weight of one bale in kilograms",
        },
        "action": "update_only",  # never inserts new product_fabric_lines rows automatically
    },
    "pieces_receipts": {
        "description": (
            "Append-only log of finished-piece receipts (non-PO). Each row is an "
            "event of 'we received N pieces of fabric X under category Y on date D'. "
            "Use this when the Excel looks like a daily/weekly receipt log."
        ),
        "match_key": ["category", "fabric_code"],  # used to resolve product_fabric_line_id
        "writable_columns": {
            "pieces": "int > 0 — how many pieces received",
            "received_at": "ISO date (YYYY-MM-DD), defaults to today if blank",
            "mill_name": "string — the source mill / contractor (optional)",
            "notes": "string (optional)",
        },
        "action": "insert",
    },
    "fabric_meter_receipts": {
        "description": (
            "Append-only log of fabric (meter) receipts. Each row is an event of "
            "'we received N meters of fabric X under category Y on date D'. Use "
            "this when the Excel records fabric coming in from a mill."
        ),
        "match_key": ["category", "fabric_code"],
        "writable_columns": {
            "meters": "decimal > 0 — how many meters received",
            "received_at": "ISO date (YYYY-MM-DD), defaults to today if blank",
            "mill_name": "string — the source mill (optional)",
            "notes": "string (optional)",
        },
        "action": "insert",
    },
}


@dataclass
class ExcelSample:
    sheet_name: str
    headers: list[str]
    rows: list[list[Any]]  # first N data rows
    total_rows: int


@dataclass
class ImportPlan:
    target_table: TargetTable
    column_mapping: dict[str, str]  # db_column -> excel_header
    match_columns: dict[str, str]   # "category" -> excel_header (one row only), used to find the line
    action: ActionKind
    confidence: float
    reasoning: str
    warnings: list[str]


def read_excel_sample(file_bytes: bytes, sample_rows: int = 10) -> ExcelSample:
    """Parse the uploaded file and return headers + a small sample. No DB I/O."""
    from openpyxl import load_workbook  # local import — heavy module, only paid when used

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        raise ValueError(f"Could not read the Excel file: {e}") from e

    ws = wb.active
    if ws is None:
        raise ValueError("The Excel file has no active sheet.")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ValueError("The sheet is empty.") from None

    headers = [str(c).strip() if c is not None else "" for c in header_row]

    rows: list[list[Any]] = []
    total = 0
    for raw in rows_iter:
        if all(v is None or str(v).strip() == "" for v in raw):
            continue
        total += 1
        if len(rows) < sample_rows:
            rows.append([_jsonable(v) for v in raw])

    return ExcelSample(sheet_name=ws.title, headers=headers, rows=rows, total_rows=total)


def _jsonable(value: Any) -> Any:
    """openpyxl returns datetime/Decimal which aren't JSON-serializable directly."""
    if value is None:
        return None
    if isinstance(value, (int, float, bool, str)):
        return value
    return str(value)


def _build_prompt(sample: ExcelSample) -> str:
    catalog_lines: list[str] = []
    for name, info in TABLE_CATALOG.items():
        cols = ", ".join(f"{k} ({v})" for k, v in info["writable_columns"].items())
        catalog_lines.append(
            f"Table `{name}`: {info['description']}\n"
            f"  Match key: ({', '.join(info['match_key'])})\n"
            f"  Writable columns: {cols}\n"
            f"  Action: {info['action']}"
        )

    sample_block = {
        "sheet": sample.sheet_name,
        "headers": sample.headers,
        "first_rows": sample.rows,
        "approx_total_rows": sample.total_rows,
    }

    return f"""You are an Excel-import assistant for a garment factory MRP.

The owner uploaded an Excel file. Pick the target DB table that best fits the
data and map each useful Excel column to a writable DB column.

AVAILABLE TARGET TABLES
{chr(10).join(catalog_lines)}

NOTES
- "category" in the database refers to a price-tier product name like "109",
  "199-PKD", "299", "399", "499". The Excel might call this column "Category",
  "MRP", "Price Tier", or similar.
- "fabric_code" is the variant name like "ASSORTED", "MISTY", "GARDEN-BLOOM".
  The Excel might call it "Fabric", "Item", "Variant", "Design".
- The Excel rows for `pieces_receipts` / `fabric_meter_receipts` describe
  *events* — each row inserts a new receipt and atomically updates the running
  inventory counter.
- For `product_fabric_lines`, only update columns the user actually provided.
  Empty cells must NOT overwrite existing values.
- Status synonyms: "ok" -> ok, "extra" -> extra, "in stock"/"instock" -> in_stock,
  "nil"/"zero" -> nil, "short"/"shortage"/"deficit" -> short.
- If you cannot confidently pick a target, set confidence < 0.5 and explain.
- The `match_columns` map says which Excel headers carry the (category, fabric_code)
  values used to find the right `product_fabric_lines` row.

EXCEL SAMPLE
{json.dumps(sample_block, ensure_ascii=False, indent=2)}

Return ONLY the JSON object that matches the response schema. No prose, no markdown."""


_RESPONSE_SCHEMA: dict[str, Any] = {
    "type": "object",
    "required": [
        "target_table",
        "column_mapping",
        "match_columns",
        "action",
        "confidence",
        "reasoning",
        "warnings",
    ],
    "properties": {
        "target_table": {
            "type": "string",
            "enum": ["product_fabric_lines", "pieces_receipts", "fabric_meter_receipts"],
        },
        "column_mapping": {
            "type": "object",
            "description": "DB column name -> Excel header (or empty string if not present)",
            "additionalProperties": {"type": "string"},
        },
        "match_columns": {
            "type": "object",
            "description": "Database key column -> Excel header that carries its value",
            "additionalProperties": {"type": "string"},
        },
        "action": {"type": "string", "enum": ["upsert", "insert", "update_only"]},
        "confidence": {"type": "number"},
        "reasoning": {"type": "string"},
        "warnings": {
            "type": "array",
            "items": {"type": "string"},
        },
    },
}


async def analyze_with_gemini(sample: ExcelSample) -> ImportPlan:
    """Ask Gemini to produce a structured import plan. Single one-shot call."""
    client = get_client()
    prompt = _build_prompt(sample)

    config = types.GenerateContentConfig(
        temperature=0.1,
        response_mime_type="application/json",
        response_schema=_RESPONSE_SCHEMA,
    )

    response = client.models.generate_content(
        model=get_model(),
        contents=prompt,
        config=config,
    )

    text = (getattr(response, "text", None) or "").strip()
    if not text:
        raise ValueError("Gemini returned an empty response.")

    try:
        data = json.loads(text)
    except json.JSONDecodeError as e:
        raise ValueError(f"Gemini returned non-JSON output: {text[:200]}") from e

    target = data.get("target_table")
    if target not in TABLE_CATALOG:
        raise ValueError(f"Gemini picked an unknown target table: {target}")

    return ImportPlan(
        target_table=target,
        column_mapping={k: v for k, v in (data.get("column_mapping") or {}).items() if v},
        match_columns={k: v for k, v in (data.get("match_columns") or {}).items() if v},
        action=data.get("action") or TABLE_CATALOG[target]["action"],
        confidence=float(data.get("confidence") or 0),
        reasoning=str(data.get("reasoning") or ""),
        warnings=list(data.get("warnings") or []),
    )


# ---------------------------------------------------------------------------
# Coercion helpers used by both preview-row generation and commit execution.
# ---------------------------------------------------------------------------

_STATUS_SYNONYMS = {
    "ok": "ok",
    "extra": "extra",
    "in_stock": "in_stock",
    "instock": "in_stock",
    "stock": "in_stock",
    "nil": "nil",
    "zero": "nil",
    "short": "short",
    "deficit": "short",
    "shortage": "short",
    "unknown": "unknown",
    "": "unknown",
}


def coerce_int(raw: Any) -> int | None:
    if raw is None or raw == "":
        return None
    try:
        return int(float(str(raw).replace(",", "")))
    except (TypeError, ValueError):
        return None


def coerce_decimal(raw: Any) -> Decimal | None:
    if raw is None or raw == "":
        return None
    try:
        return Decimal(str(raw).replace(",", ""))
    except (TypeError, ValueError, InvalidOperation):
        return None


def coerce_status(raw: Any) -> str | None:
    if raw is None or str(raw).strip() == "":
        return None
    norm = str(raw).strip().lower().replace(" ", "_").replace("-", "_")
    return _STATUS_SYNONYMS.get(norm)


def coerce_str(raw: Any) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    return s or None
